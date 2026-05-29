"""Exportação dos resultados em TXT, CSV, JSON e Excel."""

import csv
import json
from pathlib import Path

from consultas_cnmp.scraper import Processo

_CAMPOS = [
    "numero",
    "localizacao",
    "data_distribuicao",
    "relator",
    "classe_processual",
    "objeto",
]


def exportar(processos: list[Processo], destino: Path, formatos: list[str]) -> list[Path]:
    """Exporta a lista de processos nos formatos solicitados.

    Retorna a lista de arquivos gerados.
    """
    destino.mkdir(parents=True, exist_ok=True)
    arquivos = []

    dispatch = {
        "txt": _exportar_txt,
        "csv": _exportar_csv,
        "json": _exportar_json,
        "excel": _exportar_excel,
    }

    for fmt in formatos:
        fn = dispatch.get(fmt.lower())
        if fn is None:
            print(f"  Formato desconhecido ignorado: {fmt}")
            continue
        caminho = fn(processos, destino)
        arquivos.append(caminho)
        print(f"  [{fmt.upper()}] {caminho}")

    return arquivos


def _exportar_txt(processos: list[Processo], destino: Path) -> Path:
    caminho = destino / "processos.txt"
    caminho.write_text(
        "\n".join(p.numero for p in processos), encoding="utf-8"
    )
    return caminho


def _exportar_csv(processos: list[Processo], destino: Path) -> Path:
    caminho = destino / "processos.csv"
    with caminho.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_CAMPOS)
        writer.writeheader()
        for p in processos:
            writer.writerow({c: getattr(p, c) for c in _CAMPOS})
    return caminho


def _exportar_json(processos: list[Processo], destino: Path) -> Path:
    caminho = destino / "processos.json"
    dados = [{c: getattr(p, c) for c in _CAMPOS} for p in processos]
    caminho.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
    return caminho


def _exportar_excel(processos: list[Processo], destino: Path) -> Path:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Instale openpyxl: pip install openpyxl")

    caminho = destino / "processos.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos CNMP"

    cabecalhos = [
        "Número", "Localização Atual", "Data Distribuição",
        "Relator", "Classe Processual", "Objeto",
    ]
    ws.append(cabecalhos)

    # Estilo do cabeçalho
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for p in processos:
        ws.append([getattr(p, c) for c in _CAMPOS])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(caminho)
    return caminho
